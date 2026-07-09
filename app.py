import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO
from pathlib import Path
import base64
from openpyxl import load_workbook
import numpy as np
import textwrap

from pdf_generator import P3_PALETA_GRAFICO
from pdf_generator import generar_pdf

# =====================================================================
# CONFIGURACIÓN GLOBAL (ESTILOS, COLORES, FUENTES Y TAMAÑOS)
# =====================================================================

# 1. COLORES INSTITUCIONALES Y PALETAS
COLOR_PRIMARIO = "#30A907"      # Verde institucional
COLOR_SECUNDARIO = "#013051"    # Azul institucional
COLOR_TEXTO = "#013051"         # Color del texto en gráficos
COLOR_GRILLA = "#E0E0E0"        # Color de las líneas de cuadrícula

# Paleta completa para gráficos de pastel o múltiples barras
PALETA_COMPLETA = [
    "#4472C4", "#5B9BD5", "#A5A5A5", "#70AD47", "#255E91", 
    "#B7B7B7", "#9DC3E6", "#8FAADC", "#D9E1F2", "#264478",
    "#30A907", "#636363"
]

# Paletas específicas (Mantiene la identidad de tu app original)
PALETA_COMERCIO_1 = ["#5b9bd5", "#a5a5a5"]
PALETA_COMERCIO_2 = ["#4472c4", "#9dc3e6"]
PALETA_SERVICIO = ["#5b9bd5", "#a5a5a5", "#4472c4", "#255e91", "#636363"]

# 2. TAMAÑOS DE GRÁFICOS (Ancho, Alto)
SIZE_GRAFICO_BARRAS = (9, 6)
SIZE_GRAFICO_PASTEL = (7, 7)
SIZE_GRAFICO_PASTEL_PEQUENO = (6, 6)
SIZE_GRAFICO_PASTEL_MEDIANO = (5, 5)
SIZE_GRAFICO_LINEAL = (8, 5)
SIZE_GRAFICO_ATENCION = (8, 4)
SIZE_GRAFICO_SERVICIO_POLICIAL = (11, 6)
DPI_ESTANDAR = 300 # o 200 si prefieres menos peso

# 3. FUENTES Y TEXTOS PARA GRÁFICOS
FONT_SIZE_TITULOS = 18
FONT_SIZE_EJES = 14
FONT_SIZE_ETIQUETAS = 12
FONT_SIZE_PORCENTAJES_PASTEL = 25
FONT_SIZE_BARRAS_VALORES = 10
FONT_SIZE_TEXTO_DESTACADO = 20

# Configurar Matplotlib para usar estas fuentes globalmente
plt.rcParams.update({
    'font.size': FONT_SIZE_EJES,
    'axes.titlesize': FONT_SIZE_TITULOS,
    'axes.labelsize': FONT_SIZE_EJES,
    'xtick.labelsize': FONT_SIZE_EJES,
    'ytick.labelsize': FONT_SIZE_EJES,
    'text.color': COLOR_TEXTO,
    'axes.labelcolor': COLOR_TEXTO,
    'xtick.color': COLOR_TEXTO,
    'ytick.color': COLOR_TEXTO
})

# 4. FUENTES DE TABLAS (Variables para enviar a pdf_generator.py)
TABLA_FONT_FAMILY = "Helvetica"
TABLA_FONT_SIZE_HEADER = 12
TABLA_FONT_SIZE_BODY = 10
TABLA_COLOR_HEADER = COLOR_SECUNDARIO
TABLA_COLOR_TEXTO = "#000000"

# ================= STREAMLIT =================
st.set_page_config(page_title="Generador de PDF", layout="centered")
st.title("Generador de Informes SS")

# ================= RUTAS =====================
BASE_DIR = Path(__file__).resolve().parent
ASSETS_DIR = BASE_DIR / "assets"
ASSETS_DIR.mkdir(parents=True, exist_ok=True)

# ================== UTILIDADES ===============
def limpiar_series(labels, values):
    df = pd.DataFrame({"label": labels, "value": values})
    df["value"] = (
        df["value"]
        .astype(str)
        .str.replace("%", "", regex=False)
        .str.replace(",", ".", regex=False)
        .str.strip()
    )
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df = df.dropna()
    return df["label"].astype(str), df["value"]

def crear_grafico(labels, values):
    fig, ax = plt.subplots(figsize=SIZE_GRAFICO_BARRAS)
    ax.bar(labels, values, color=COLOR_PRIMARIO)
    ax.set_ylim(0, 100)
    ax.set_ylabel("%", fontsize=FONT_SIZE_EJES)
    ax.tick_params(axis='x', labelsize=FONT_SIZE_EJES)
    ax.tick_params(axis='y', labelsize=FONT_SIZE_EJES)
    ax.set_title("Relación por distrito", fontsize=FONT_SIZE_TITULOS)
    
    for i, v in enumerate(values):
        ax.text(i, v + 1, f"{v:.0f}%", ha="center", fontsize=FONT_SIZE_ETIQUETAS)
    
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=DPI_ESTANDAR)
    plt.close(fig)
    buf.seek(0)
    return buf

def seguro_int(valor, default=0):
    if pd.isna(valor):
        return default
    try:
        return int(float(valor))
    except:
        return default

# ================= APP =================
archivo = st.file_uploader(
    "Aquí suba o arrastre el ENGINE de la Delegación correspondiente",
    type=["xlsx"]
)

if archivo:
    try:
        # Lectura del excel
        wb = load_workbook(archivo, data_only=True)
        ws = wb["Hoja1"]

        # ================= CORRESPONSABLES DESDE EXCEL ============
        columnas_corresponsable = ["J", "P", "V", "AB", "AH", "AN", "AT", "AZ", "BF", "BL", "BR", "BX"]
        corresponsables_excel = []
        for col in columnas_corresponsable:
            valor = ws[f"{col}246"].value
            if valor:
                corresponsables_excel.append(str(valor).strip())
            else:
                corresponsables_excel.append("")
                
        df = pd.DataFrame(ws.values)

        # ================= DATOS BASE ========
        delegacion = str(df.iloc[1, 1]) #B2
        codigo = str(df.iloc[2, 1]) #B3

        # ================= TABLA PARTICIPACIÓN =================
        tabla_df = df.iloc[6:23, 0:3].dropna(how="all")
        tabla_df = tabla_df[~(tabla_df.iloc[:, 1:].fillna(0).astype(str) == "0").all(axis=1)]

        def formatear(v):
            if isinstance(v, (int, float)):
                if 0 < v < 1:
                    return f"{v*100:.0f}%"
                return f"{v:.0f}"
            return str(v)

        tabla_participacion = [
            [formatear(c) for c in fila]
            for fila in tabla_df.fillna("").values.tolist()
        ]

        # ================= GRÁFICO RELACIÓN =================
        rel_labels = df.iloc[7:11, 6].astype(str)
        rel_base_values = pd.to_numeric(df.iloc[7:11, 8], errors="coerce")
        rel_percent_labels = df.iloc[7:11, 7].astype(str).str.replace(",", "", regex=False)

        mask = rel_base_values.notna()
        rel_labels = rel_labels[mask]
        rel_base_values = rel_base_values[mask]
        rel_percent_labels = rel_percent_labels[mask]

        fig, ax = plt.subplots(figsize=SIZE_GRAFICO_BARRAS)
        ax.bar(rel_labels, rel_base_values, color=COLOR_PRIMARIO)
        ax.set_ylabel("Cantidad")
        ax.margins(y=0.1)

        for spine in ax.spines.values():
            spine.set_visible(False)
        ax.tick_params(left=False, bottom=False)
        ax.set_facecolor("none")
        fig.patch.set_alpha(0)

        for i in range(len(rel_percent_labels)):
            porcentaje = float(rel_percent_labels.iloc[i]) * 100
            ax.text(
                i,
                rel_base_values.iloc[i],
                f"{porcentaje:.2f}%",
                ha="center",
                va="bottom",
                fontsize=FONT_SIZE_ETIQUETAS,
                color=COLOR_TEXTO
            )

        buf_rel = BytesIO()
        fig.savefig(buf_rel, format="png", bbox_inches="tight", pad_inches=0, dpi=DPI_ESTANDAR, transparent=True)
        plt.close(fig)
        buf_rel.seek(0)
        grafico_rel_path = BASE_DIR / "grafico_relacion.png"
        with open(grafico_rel_path, "wb") as f:
            f.write(buf_rel.getbuffer())

        # ================= PARTICIPACIÓN POR EDAD =================
        edad_labels = df.iloc[28:33, 0].astype(str)
        edad_percent_values = (
            df.iloc[28:33, 1]
            .astype(str)
            .str.replace("%", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        edad_percent_values = pd.to_numeric(edad_percent_values, errors="coerce")
        mask = edad_percent_values.notna()
        edad_labels = edad_labels[mask]
        edad_percent_values = edad_percent_values[mask]

        fig_edad, ax_edad = plt.subplots(figsize=SIZE_GRAFICO_PASTEL)
        wedges, texts, autotexts = ax_edad.pie(
            edad_percent_values,
            labels=None,
            autopct=lambda p: f"{p:.0f}%",
            pctdistance=0.65,
            labeldistance=1.15,
            startangle=90,
            colors=PALETA_COMPLETA[:5],
            textprops={"fontsize": FONT_SIZE_ETIQUETAS}
        )
        ax_edad.axis("equal")
        for autotext in autotexts:
            autotext.set_fontsize(FONT_SIZE_PORCENTAJES_PASTEL)
        
        ax_edad.set_facecolor("none")
        fig_edad.patch.set_alpha(0)

        buf_edad = BytesIO()
        fig_edad.savefig(buf_edad, format="png", dpi=DPI_ESTANDAR, transparent=True)
        plt.close(fig_edad)
        buf_edad.seek(0)
        grafico_edad_path = BASE_DIR / "grafico_participacion_edad.png"
        with open(grafico_edad_path, "wb") as f:
            f.write(buf_edad.getbuffer())

        # Tabla Edad
        tabla_edad_df = df.iloc[28:33, 0:2].copy()
        tabla_edad_df.iloc[:, 1] = tabla_edad_df.iloc[:, 1].apply(
            lambda x: f"{x*100:.0f}%" if isinstance(x, (int, float)) else x
        )
        tabla_edad = tabla_edad_df.fillna("").values.tolist()

        # ================= ESCOLARIDAD =================
        escolaridad_labels = df.iloc[38:46, 0].astype(str)
        escolaridad_percent_values = (
            df.iloc[38:46, 1]
            .astype(str)
            .str.replace("%", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        escolaridad_percent_values = pd.to_numeric(escolaridad_percent_values, errors="coerce")
        mask = escolaridad_percent_values.notna()
        escolaridad_labels = escolaridad_labels[mask]
        escolaridad_percent_values = escolaridad_percent_values[mask]

        fig_esco, ax_esco = plt.subplots(figsize=SIZE_GRAFICO_PASTEL)
        wedges, texts, autotexts = ax_esco.pie(
            escolaridad_percent_values,
            labels=None,
            autopct=lambda p: f"{p:.0f}%",
            pctdistance=0.65,
            labeldistance=1.15,
            startangle=90,
            colors=PALETA_COMPLETA[:8],
            textprops={"fontsize": FONT_SIZE_ETIQUETAS}
        )
        ax_esco.axis("equal")
        for autotext in autotexts:
            autotext.set_fontsize(FONT_SIZE_PORCENTAJES_PASTEL)
        ax_esco.set_facecolor("none")
        fig_esco.patch.set_alpha(0)

        buf_esco = BytesIO()
        fig_esco.savefig(buf_esco, format="png", dpi=DPI_ESTANDAR, transparent=True)
        plt.close(fig_esco)
        buf_esco.seek(0)
        grafico_escolaridad_path = BASE_DIR / "grafico_participacion_escolaridad.png"
        with open(grafico_escolaridad_path, "wb") as f:
            f.write(buf_esco.getbuffer())

        tabla_escolaridad_df = df.iloc[38:46, 0:2].copy()
        tabla_escolaridad_df.iloc[:, 1] = tabla_escolaridad_df.iloc[:, 1].apply(
            lambda x: f"{x*100:.0f}%" if isinstance(x, (int, float)) else x
        )
        tabla_escolaridad = tabla_escolaridad_df.fillna("").values.tolist()

        # ================= PARTICIPACIÓN POR GÉNERO =================
        genero_labels = df.iloc[51:54, 0].astype(str)
        genero_percent_values = (
            df.iloc[51:54, 1]
            .astype(str)
            .str.replace("%", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        genero_percent_values = pd.to_numeric(genero_percent_values, errors="coerce")
        mask = genero_percent_values.notna()
        genero_labels = genero_labels[mask]
        genero_percent_values = genero_percent_values[mask]

        fig_gen, ax_gen = plt.subplots(figsize=SIZE_GRAFICO_PASTEL)
        wedges, texts, autotexts = ax_gen.pie(
            genero_percent_values,
            labels=None,
            autopct=lambda p: f"{p:.0f}%",
            pctdistance=0.65,
            labeldistance=1.15,
            startangle=90,
            colors=PALETA_COMPLETA[:3],
            textprops={"fontsize": FONT_SIZE_ETIQUETAS}
        )
        ax_gen.axis("equal")
        for autotext in autotexts:
            autotext.set_fontsize(FONT_SIZE_PORCENTAJES_PASTEL)
        ax_gen.set_facecolor("none")
        fig_gen.patch.set_alpha(0)

        buf_gen = BytesIO()
        fig_gen.savefig(buf_gen, format="png", dpi=DPI_ESTANDAR, transparent=True)
        plt.close(fig_gen)
        buf_gen.seek(0)
        grafico_genero_path = BASE_DIR / "grafico_participacion_genero.png"
        with open(grafico_genero_path, "wb") as f:
            f.write(buf_gen.getbuffer())

        tabla_genero_df = df.iloc[51:54, 0:2].copy()
        tabla_genero_df.iloc[:, 1] = tabla_genero_df.iloc[:, 1].apply(
            lambda x: f"{x*100:.0f}%" if isinstance(x, (int, float)) else x
        )
        tabla_genero = tabla_genero_df.fillna("").values.tolist()

        # ================= TABLAS ENCUESTAS =================
        tabla_encuesta_comunidad_df = df.iloc[58:60, 0:4].copy()
        tabla_encuesta_comunidad = tabla_encuesta_comunidad_df.fillna("").values.tolist()

        tabla_otras_encuestas_df = df.iloc[62:65, 6:10].copy()
        tabla_otras_encuestas = tabla_otras_encuestas_df.fillna("").values.tolist()

        datos_pagina_8 = {
            "encuesta_comunidad": seguro_int(df.iloc[82, 1]),
            "encuesta_policial": seguro_int(df.iloc[83, 1]),
            "encuesta_comercio": seguro_int(df.iloc[84, 1]),
            "estadistica": seguro_int(df.iloc[86, 2]),
            "total_datos": seguro_int(df.iloc[87, 1])
        }

        datos_pagina_9 = {
            "lado_izquierdo": str(df.iloc[92, 0]),
            "derecha_superior": str(df.iloc[92, 1]),
            "derecha_inferior": str(df.iloc[92, 2])
        }

        # ================= PARETO / DELITOS =================
        tabla_delitos_raw = df.iloc[96:117, 1]
        tabla_delitos = [
            [str(v)] for v in tabla_delitos_raw
            if pd.notna(v) and str(v) != "0"
        ]

        tabla_riesgos_raw = df.iloc[96:117, 2]
        tabla_riesgos = [
            [str(v)] for v in tabla_riesgos_raw
            if pd.notna(v) and str(v) != "0"
        ]

        valor_delitos = df.iloc[118, 1]
        valor_riesgos = df.iloc[118, 2]
        porcentaje_delitos = f"{valor_delitos * 100:.2f}%" if pd.notna(valor_delitos) else "0.00%"
        porcentaje_riesgos = f"{valor_riesgos * 100:.2f}%" if pd.notna(valor_riesgos) else "0.00%"

        cantidad_delitos = int(pd.to_numeric(df.iloc[117, 1], errors="coerce"))
        cantidad_riesgos = int(pd.to_numeric(df.iloc[117, 2], errors="coerce"))

        # ================= MICMAC =================
        def limpiar_lista(col):
            return [[str(v)] for v in col if pd.notna(v) and str(v).strip() != ""]

        micmac_poder = limpiar_lista(df.iloc[123:140, 1])
        micmac_conflicto = limpiar_lista(df.iloc[123:140, 2])
        micmac_autonomas = limpiar_lista(df.iloc[123:140, 3])
        micmac_resultados = limpiar_lista(df.iloc[123:140, 4])

        def limpiar_lista_simple(col):
            return [[str(v)] for v in col if pd.notna(v) and str(v).strip() != ""]

        tabla_riesgos_micmac2 = limpiar_lista_simple(df.iloc[123:140, 10])
        tabla_delitos_micmac2 = limpiar_lista_simple(df.iloc[123:140, 11])

        cantidad_problematicas = int(df.iloc[140, 12])
        riesgos_total = int(df.iloc[140, 10])
        delitos_total = int(df.iloc[140, 11])

        causas_identificadas = int(df.iloc[117, 3])
        factores_micmac = int(df.iloc[140, 12])
        triangulo_directa = int(df.iloc[146, 0])
        triangulo_sociocultural = int(df.iloc[146, 1])
        triangulo_estructural = int(df.iloc[146, 2])

        # ================= INSTITUCIONES =================
        tabla_instituciones_df = df.iloc[149:159, 1:3].copy()
        tabla_instituciones_df = tabla_instituciones_df.dropna(how="all")
        tabla_instituciones = []
        for _, row in tabla_instituciones_df.iterrows():
            col1 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            col2 = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if col1 or col2:
                tabla_instituciones.append([col1, col2])

        # ================= ESTADÍSTICA =================
        df_grafico_denuncias = df.iloc[165:176, [0, 2]].copy()
        df_grafico_denuncias.columns = ["categoria", "porcentaje"]
        df_grafico_denuncias["porcentaje"] = (
            df_grafico_denuncias["porcentaje"]
            .astype(str)
            .str.replace("%", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df_grafico_denuncias["porcentaje"] = pd.to_numeric(df_grafico_denuncias["porcentaje"], errors="coerce")
        df_grafico_denuncias = df_grafico_denuncias.dropna()

        df_tabla_denuncias = df.iloc[165:176, [0, 1]].copy()
        df_tabla_denuncias.columns = ["categoria", "cantidad"]
        df_tabla_denuncias = df_tabla_denuncias.dropna(how="all")
        tabla_denuncias = [
            [str(row["categoria"]), str(int(row["cantidad"]))]
            for _, row in df_tabla_denuncias.iterrows()
            if pd.notna(row["categoria"]) and pd.notna(row["cantidad"])
        ]
        total_denuncias = int(df.iloc[177, 1])

        def generar_grafico_denuncias(df):
            fig, ax = plt.subplots(figsize=SIZE_GRAFICO_PASTEL_PEQUENO)
            ax.pie(
                df["porcentaje"],
                labels=df["categoria"],
                autopct="%1.0f%%",
                startangle=90,
                colors=PALETA_COMPLETA[:len(df)],
                textprops={'fontsize': FONT_SIZE_ETIQUETAS}
            )
            ax.axis("equal")
            plt.tight_layout()
            plt.savefig(ASSETS_DIR / "grafico_denuncias.png", dpi=DPI_ESTANDAR)
            plt.close()

        generar_grafico_denuncias(df_grafico_denuncias)

        # ================= HORARIOS =================
        df_grafico_horario = df.iloc[179:188, [0, 2]].copy()
        df_grafico_horario.columns = ["horario", "porcentaje"]
        df_grafico_horario["porcentaje"] = (
            df_grafico_horario["porcentaje"]
            .astype(str)
            .str.replace("%", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df_grafico_horario["porcentaje"] = pd.to_numeric(df_grafico_horario["porcentaje"], errors="coerce")
        df_grafico_horario = df_grafico_horario.dropna()

        df_tabla_horario = df.iloc[179:188, [0, 1]].copy()
        df_tabla_horario.columns = ["horario", "cantidad"]
        tabla_horario = [
            [str(row["horario"]), str(row["cantidad"])]
            for _, row in df_tabla_horario.iterrows()
            if pd.notna(row["horario"])
        ]

        total_am = df.iloc[179, 3]
        total_pm = df.iloc[179, 4]

        def formatear_porcentaje(valor):
            if pd.notna(valor):
                try:
                    return f"{float(valor) * 100:.2f}%"
                except:
                    return str(valor)
            return ""

        total_am = formatear_porcentaje(total_am)
        total_pm = formatear_porcentaje(total_pm)

        encabezados = df.iloc[178, 0:17].copy()
        tabla_datos = df.iloc[179:188, 0:17].copy()
        columnas_a_eliminar = [3, 4]
        encabezados = encabezados.drop(encabezados.index[columnas_a_eliminar])
        tabla_datos = tabla_datos.drop(tabla_datos.columns[columnas_a_eliminar], axis=1)

        tabla_datos.iloc[:, 2] = tabla_datos.iloc[:, 2].apply(formatear_porcentaje)
        tabla_datos = tabla_datos.loc[:, ~(tabla_datos.isna().all())]
        
        tabla_horario_distrito = [
            encabezados.loc[tabla_datos.columns].fillna("").astype(str).tolist()
        ]
        tabla_horario_distrito += tabla_datos.fillna("").astype(str).values.tolist()

        def generar_grafico_horario(df):
            fig, ax = plt.subplots(figsize=SIZE_GRAFICO_PASTEL_PEQUENO)
            ax.pie(
                df["porcentaje"],
                labels=df["horario"],
                autopct="%1.0f%%",
                startangle=90,
                colors=PALETA_COMPLETA[:len(df)],
                textprops={'fontsize': FONT_SIZE_ETIQUETAS}
            )
            ax.axis("equal")
            plt.tight_layout()
            plt.savefig(ASSETS_DIR / "grafico_horario.png", dpi=DPI_ESTANDAR, transparent=True)
            plt.close()

        generar_grafico_horario(df_grafico_horario)

        # ================= GRÁFICO BARRAS PÁGINA 14 =================
        df_grafico_p14 = df.iloc[195:204, [0, 1]].copy()
        df_grafico_p14.columns = ["categoria", "valor"]
        df_grafico_p14 = df_grafico_p14.dropna()
        df_grafico_p14["valor"] = pd.to_numeric(df_grafico_p14["valor"], errors="coerce")
        df_grafico_p14 = df_grafico_p14.dropna()

        def generar_grafico_p14(df):
            fig, ax = plt.subplots(figsize=SIZE_GRAFICO_LINEAL)
            barras = ax.bar(df["categoria"], df["valor"], color=COLOR_PRIMARIO)
            ax.set_ylabel("")
            ax.set_xlabel("")
            ax.set_title("")
            ax.tick_params(axis="x", rotation=45, labelsize=FONT_SIZE_EJES)
            
            for spine in ax.spines.values():
                spine.set_visible(False)
            ax.tick_params(left=False, bottom=False)

            for bar in barras:
                height = bar.get_height()
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    height,
                    f"{int(height)}",
                    ha="center",
                    va="bottom",
                    fontsize=FONT_SIZE_BARRAS_VALORES,
                    color=COLOR_TEXTO
                )
            plt.tight_layout()
            plt.savefig(ASSETS_DIR / "grafico_p14.png", dpi=DPI_ESTANDAR, transparent=True)
            plt.close()

        generar_grafico_p14(df_grafico_p14)

        # ================= TABLA P14 =================
        encabezados_p14 = df.iloc[206, 0:9].copy()
        tabla_p14_raw = df.iloc[207:219, 0:9].copy()
        tabla_p14_raw = tabla_p14_raw.dropna(how="all")
        
        tabla_p14 = []
        tabla_p14.append(encabezados_p14.fillna("").astype(str).tolist())
        
        for _, row in tabla_p14_raw.iterrows():
            distrito = row.iloc[0]
            if pd.isna(distrito):
                continue
            distrito_str = str(distrito).strip()
            if distrito_str == "0":
                continue
            
            frecuencias = row.iloc[1:]
            frecuencias_numericas = pd.to_numeric(frecuencias, errors="coerce").fillna(0)
            if (frecuencias_numericas == 0).all():
                continue
                
            fila = []
            for cell in row:
                if pd.isna(cell):
                    fila.append("")
                else:
                    if isinstance(cell, (int, float)):
                        fila.append(str(int(cell)))
                    else:
                        fila.append(str(cell))
            tabla_p14.append(fila)

        # ================= GRÁFICO LINEAL PÁGINA 15 =================
        df_grafico_p15 = df.iloc[222:229, 0:3].copy()
        df_grafico_p15.columns = ["dia", "frecuencia", "porcentaje"]
        df_grafico_p15 = df_grafico_p15.dropna(how="all")
        df_grafico_p15["frecuencia"] = pd.to_numeric(df_grafico_p15["frecuencia"], errors="coerce")
        df_grafico_p15["porcentaje"] = (
            df_grafico_p15["porcentaje"]
            .astype(str)
            .str.replace("%", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df_grafico_p15["porcentaje"] = pd.to_numeric(df_grafico_p15["porcentaje"], errors="coerce")
        df_grafico_p15 = df_grafico_p15.dropna()

        def generar_grafico_p15(df):
            fig, ax = plt.subplots(figsize=SIZE_GRAFICO_LINEAL)
            x_vals = range(len(df))
            
            ax.plot(x_vals, df["frecuencia"], marker="o", linewidth=3, markersize=8, color=COLOR_SECUNDARIO)
            ax.fill_between(x_vals, df["frecuencia"], color=COLOR_PRIMARIO, alpha=0.15)
            ax.grid(axis='y', linestyle='--', alpha=0.4, color=COLOR_GRILLA)

            for i, row in enumerate(df.itertuples()):
                ax.text(
                    i, row.frecuencia, f"{int(row.frecuencia)}",
                    ha="center", va="bottom", fontsize=FONT_SIZE_BARRAS_VALORES,
                    fontweight="bold", color=COLOR_TEXTO
                )

            ax.set_ylim(0, df["frecuencia"].max() * 1.25)
            offset = df["frecuencia"].max() * 0.08
            
            for i, row in enumerate(df.itertuples()):
                ax.text(
                    i, row.frecuencia - offset, f"{row.porcentaje * 100:.2f}%",
                    ha="center", va="top", fontsize=FONT_SIZE_BARRAS_VALORES, color=COLOR_TEXTO
                )

            ax.set_xticks(x_vals)
            ax.set_xticklabels(df["dia"], fontsize=FONT_SIZE_EJES)
            
            for spine in ax.spines.values():
                spine.set_visible(False)
            ax.tick_params(left=False, bottom=False)
            ax.set_ylabel("")
            ax.set_xlabel("")
            ax.set_title("")
            
            plt.tight_layout()
            plt.savefig(ASSETS_DIR / "grafico_p15.png", dpi=DPI_ESTANDAR, transparent=True)
            plt.close()

        generar_grafico_p15(df_grafico_p15)

        # ================= TABLA P15 =================
        dias_p15 = df.iloc[222:229, 0].copy().astype(str).tolist()
        distritos_p15 = df.iloc[221, 3:15].copy().astype(str).tolist()
        valores_p15_raw = df.iloc[222:229, 3:15].copy()
        
        tabla_p15 = []
        tabla_p15.append(["Distrito"] + dias_p15)
        
        for i, distrito in enumerate(distritos_p15):
            fila = [distrito]
            for valor in valores_p15_raw.iloc[:, i]:
                if pd.isna(valor):
                    fila.append("")
                else:
                    fila.append(str(int(valor)) if isinstance(valor, (int, float)) else str(valor))
            if all(v == "" for v in fila[1:]):
                continue
            tabla_p15.append(fila)

        # ================= LÍNEAS DE ACCIÓN =================
        region_numero = seguro_int(df.iloc[1, 3])
        delegacion_codigo = str(df.iloc[2, 1])
        numero_delegacion = int(delegacion_codigo.replace("D-", "").replace("D", "").strip())

        def safe_int(value):
            if pd.isna(value) or str(value).strip() == "":
                return 0
            return int(value)

        lineas_municipalidad = seguro_int(df.iloc[238, 0])
        lineas_fp = seguro_int(df.iloc[238, 1])
        lineas_mixtas = safe_int(df.iloc[238, 2])
        total_lineas = seguro_int(df.iloc[238, 3])

        if lineas_mixtas == 0:
            lineas_mixtas = None

        logo_muni_path = ASSETS_DIR / "Municipalidades" / str(region_numero) / f"{numero_delegacion}.png"

        # Llenar datos de líneas
        columnas_total_porcentaje = [9, 15, 21, 27, 33, 39, 45, 51, 57, 63, 69, 75]
        columnas_causas = [5, 11, 17, 23, 29, 35, 41, 47, 53, 59, 65, 71]
        columnas_problemas = [6, 12, 18, 24, 30, 36, 42, 48, 54, 60, 66, 72]
        columnas_lider = [9, 15, 21, 27, 33, 39, 45, 51, 57, 63, 69, 75]
        columnas_acciones = [8, 14, 20, 26, 32, 38, 44, 50, 56, 62, 68, 74]
        columnas_cogestores = [8, 14, 20, 26, 32, 38, 44, 50, 56, 62, 68, 74]
        
        lineas_accion_data = []
        for i in range(int(total_lineas)):
            fila_problematicas = 241 + i
            problematicas = []
            for col in [1, 2, 3]:
                valor = df.iloc[fila_problematicas, col]
                if pd.notna(valor) and str(valor).strip() != "":
                    problematicas.append(str(valor).strip())
            
            col_causa = columnas_causas[i]
            causas = []
            for fila in range(246, 276):
                valor = df.iloc[fila, col_causa]
                if pd.notna(valor) and str(valor).strip() != "":
                    causas.append([str(valor).strip()])
                    
            col_problema = columnas_problemas[i]
            problemas = []
            for fila in range(246, 276):
                valor = df.iloc[fila, col_problema]
                if pd.notna(valor) and str(valor).strip() != "":
                    problemas.append([str(valor).strip()])
                    
            col_total = columnas_total_porcentaje[i]
            valor_total = df.iloc[241, col_total]
            if pd.notna(valor_total):
                try:
                    total_porcentaje = float(valor_total) * 100
                    total_porcentaje = f"{total_porcentaje:.2f}%"
                except:
                    total_porcentaje = "0.00%"
            else:
                total_porcentaje = "0.00%"
                
            col_lider = columnas_lider[i]
            lider_estrategico = df.iloc[245, col_lider]
            col_corresponsable = columnas_corresponsable[i]
            corresponsable = df.iloc[245, col_corresponsable]
            
            corresponsable = str(corresponsable).strip() if pd.notna(corresponsable) else ""
            lider_estrategico = str(lider_estrategico).strip() if pd.notna(lider_estrategico) else ""
            
            col_acciones = columnas_acciones[i]
            acciones = []
            for fila in range(248, 257):
                valor = df.iloc[fila, col_acciones]
                if pd.notna(valor) and str(valor).strip() != "":
                    acciones.append(str(valor).strip())
                    
            col_cogestores = columnas_cogestores[i]
            cogestores = []
            for fila in range(261, 263):
                valor = df.iloc[fila, col_cogestores]
                if pd.notna(valor) and str(valor).strip() != "":
                    partes = str(valor).split(",")
                    for p in partes:
                        if p.strip():
                            cogestores.append(p.strip())
                            
            lineas_accion_data.append({
                "numero": i + 1,
                "problematicas": problematicas,
                "causas": causas,
                "problemas_influyentes": problemas,
                "total_porcentaje": total_porcentaje,
                "lider_estrategico": lider_estrategico,
                "acciones": acciones,
                "cogestores": cogestores,
                "corresponsable": corresponsable
            })

        st.write("Cantidad de lineas detectadas:", len(lineas_accion_data))

        # ================= PERCEPCION ACTUAL =================
        df_percepcion_actual = df.iloc[283:285, 0:2].copy()
        df_percepcion_actual.columns = ["respuesta", "porcentaje"]
        df_percepcion_actual["porcentaje"] = (
            df_percepcion_actual["porcentaje"]
            .astype(str)
            .str.replace("%", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df_percepcion_actual["porcentaje"] = pd.to_numeric(df_percepcion_actual["porcentaje"], errors="coerce")
        df_percepcion_actual = df_percepcion_actual.dropna()

        def generar_grafico_percepcion_actual(df):
            fig, ax = plt.subplots(figsize=SIZE_GRAFICO_PASTEL_MEDIANO)
            COLORES_PERCEPCION = [COLOR_PRIMARIO, COLOR_SECUNDARIO, "#A5A5A5"]
            
            wedges, texts, autotexts = ax.pie(
                df["porcentaje"],
                labels=None,
                autopct=lambda p: f"{p:.2f}%",
                startangle=90,
                colors=COLORES_PERCEPCION[:len(df)],
                radius=1.25,
                pctdistance=0.62,
                textprops={"fontsize": 11, "fontweight": "bold"}
            )
            for autotext in autotexts:
                autotext.set_fontsize(FONT_SIZE_ETIQUETAS)
                autotext.set_fontweight("bold")
                autotext.set_color("white")
                
            labels = df["respuesta"].tolist()
            ax.legend(
                wedges, labels, loc="center left", bbox_to_anchor=(0.95, 0.5),
                fontsize=FONT_SIZE_BARRAS_VALORES, frameon=False
            )
            ax.axis("equal")
            plt.tight_layout()
            plt.savefig(ASSETS_DIR / "grafico_percepcion_actual.png", dpi=DPI_ESTANDAR, transparent=True, pad_inches=0.03)
            plt.close()

        generar_grafico_percepcion_actual(df_percepcion_actual)

        # ================= PERCEPCION COMPARACION =================
        df_percepcion_comparacion = df.iloc[290:293, 0:2].copy()
        df_percepcion_comparacion.columns = ["categoria", "porcentaje"]
        df_percepcion_comparacion["porcentaje"] = (
            df_percepcion_comparacion["porcentaje"]
            .astype(str)
            .str.replace("%", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df_percepcion_comparacion["porcentaje"] = pd.to_numeric(df_percepcion_comparacion["porcentaje"], errors="coerce")
        df_percepcion_comparacion = df_percepcion_comparacion.dropna()

        def generar_grafico_percepcion_comparacion(df):
            fig, ax = plt.subplots(figsize=SIZE_GRAFICO_LINEAL)
            barras = ax.bar(df["categoria"], df["porcentaje"], color=COLOR_PRIMARIO)
            ax.set_ylim(0, df["porcentaje"].max() * 1.25)
            
            for bar in barras:
                height = bar.get_height()
                ax.text(
                    bar.get_x() + bar.get_width()/2,
                    height, f"{height:.2f}%",
                    ha="center", va="bottom", fontsize=FONT_SIZE_EJES
                )
            ax.tick_params(axis="x", rotation=45, labelsize=FONT_SIZE_EJES)
            ax.tick_params(axis="y", labelsize=FONT_SIZE_EJES)
            for spine in ax.spines.values():
                spine.set_visible(False)
                
            plt.tight_layout()
            plt.savefig(ASSETS_DIR / "grafico_percepcion_comparacion.png", dpi=DPI_ESTANDAR, transparent=True)
            plt.close()

        generar_grafico_percepcion_comparacion(df_percepcion_comparacion)

        # ================= VICTIMIZACIÓN Y NO DENUNCIA =================
        df_victimizacion = df.iloc[313:316, 0:2].copy()
        df_victimizacion.columns = ["categoria", "porcentaje"]
        df_victimizacion["porcentaje"] = (
            df_victimizacion["porcentaje"]
            .astype(str)
            .str.replace("%", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df_victimizacion["porcentaje"] = pd.to_numeric(df_victimizacion["porcentaje"], errors="coerce")
        df_victimizacion = df_victimizacion.dropna()

        df_no_denuncia = df.iloc[322:330, 0:2].copy()
        df_no_denuncia.columns = ["categoria", "porcentaje"]
        df_no_denuncia["porcentaje"] = (
            df_no_denuncia["porcentaje"]
            .astype(str)
            .str.replace("%", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df_no_denuncia["porcentaje"] = pd.to_numeric(df_no_denuncia["porcentaje"], errors="coerce")
        df_no_denuncia = df_no_denuncia.dropna()

        tabla_no_denuncia_df = df.iloc[322:330, [0, 2]].copy()
        tabla_no_denuncia_df = tabla_no_denuncia_df.dropna(how="all")
        tabla_no_denuncia = []
        for _, row in tabla_no_denuncia_df.iterrows():
            categoria = row.iloc[0] if len(row) > 0 else None
            valor = row.iloc[1] if len(row) > 1 else None
            if pd.notna(categoria) and pd.notna(valor):
                tabla_no_denuncia.append([str(categoria), int(valor)])

        if not df_no_denuncia.empty and df_no_denuncia["porcentaje"].notna().any():
            fila_max = df_no_denuncia.loc[df_no_denuncia["porcentaje"].idxmax()]
            motivo_principal = str(fila_max["categoria"])
        else:
            motivo_principal = "No especificado"
            
        if df.shape[0] > 321 and df.shape[1] > 6:
            total_omitidas = seguro_int(df.iloc[321, 6])
        else:
            total_omitidas = 0

        # TABLA COMPARATIVA PERCEPCIÓN
        fila_298 = df.iloc[297, 0:13].copy()
        fila_299 = df.iloc[298, 0:13].copy()
        datos_pc = df.iloc[299:311, 0:13].copy()
        columnas_indices = [0, 2, 4, 6, 8, 10, 12]

        encabezado_superior = []
        encabezado_inferior = []
        for idx in columnas_indices:
            valor_sup = fila_298.iloc[idx]
            if pd.isna(valor_sup) and idx > 0:
                valor_sup = fila_298.iloc[idx - 1]
            encabezado_superior.append(str(valor_sup).strip() if not pd.isna(valor_sup) else "")
            valor_inf = fila_299.iloc[idx]
            encabezado_inferior.append(str(valor_inf).strip() if not pd.isna(valor_inf) else "")

        tabla_percepcion = [encabezado_superior, encabezado_inferior]
        for _, row in datos_pc.iterrows():
            fila = []
            for idx in columnas_indices:
                valor = row.iloc[idx]
                if pd.isna(valor):
                    fila.append("")
                else:
                    if isinstance(valor, float):
                        porcentaje = round(valor * 100, 1)
                        fila.append(f"{int(porcentaje)}%" if porcentaje.is_integer() else f"{porcentaje}%")
                    else:
                        fila.append(str(valor))
            if any(str(v).strip() != "" for v in fila):
                tabla_percepcion.append(fila)

        def generar_grafico_victimizacion(df, nombre_archivo):
            if df.empty: return
            FIG_WIDTH = max(18, len(df) * 3)
            
            fig, ax = plt.subplots(figsize=(FIG_WIDTH, SIZE_GRAFICO_BARRAS[1]))
            fig.patch.set_facecolor("none")

            if "no_denuncia" in nombre_archivo:
                ancho_wrap = 16
                max_lineas = 3
                bottom_space = 0.42
            else:
                ancho_wrap = 22
                max_lineas = 2
                bottom_space = 0.30

            def dividir_texto(texto):
                lineas = textwrap.wrap(str(texto), width=ancho_wrap)
                return "\n".join(lineas[:max_lineas])

            categorias_wrapped = df["categoria"].apply(dividir_texto)
            x = np.arange(len(df))
            barras = ax.bar(x, df["porcentaje"], color=COLOR_PRIMARIO, width=0.5)
            
            ax.set_ylim(0, df["porcentaje"].max() * 1.25)
            ax.set_xticks(x)
            ax.set_xticklabels(categorias_wrapped, rotation=0, ha="center", fontsize=FONT_SIZE_TEXTO_DESTACADO)
            plt.subplots_adjust(bottom=bottom_space)

            for bar in barras:
                height = bar.get_height()
                ax.text(
                    bar.get_x() + bar.get_width()/2,
                    height, f"{height:.2f}%",
                    ha="center", va="bottom", fontsize=FONT_SIZE_TEXTO_DESTACADO, color=COLOR_TEXTO
                )

            for spine in ax.spines.values():
                spine.set_visible(False)
                
            plt.tight_layout(rect=[0, 0.1, 1, 1])
            plt.savefig(ASSETS_DIR / nombre_archivo, dpi=DPI_ESTANDAR, transparent=True)
            plt.close()

        generar_grafico_victimizacion(df_victimizacion, "grafico_victimizacion.png")
        generar_grafico_victimizacion(df_no_denuncia, "grafico_no_denuncia.png")

        # ================= PÁGINA 3 =================
        def generar_grafico_horarios_percepcion(labels, valores):
            plt.figure(figsize=SIZE_GRAFICO_PASTEL_PEQUENO)
            plt.pie(
                valores, labels=None, colors=P3_PALETA_GRAFICO, 
                autopct=lambda p: f"{p:.2f}%", startangle=90, textprops={'fontsize': FONT_SIZE_EJES}
            )
            plt.axis("equal")
            ruta = ASSETS_DIR / "grafico_horarios_percepcion.png"
            plt.savefig(ruta, dpi=DPI_ESTANDAR, bbox_inches="tight")
            plt.close()
            return ruta

        def generar_grafico_armas_percepcion(labels, valores):
            plt.figure(figsize=SIZE_GRAFICO_PASTEL_PEQUENO)
            plt.pie(
                valores, labels=None, colors=P3_PALETA_GRAFICO, 
                autopct=lambda p: f"{p:.2f}%", startangle=90, textprops={'fontsize': FONT_SIZE_EJES}
            )
            plt.axis("equal")
            ruta = ASSETS_DIR / "grafico_armas_percepcion.png"
            plt.savefig(ruta, dpi=DPI_ESTANDAR, bbox_inches="tight")
            plt.close()
            return ruta

        def generar_grafico_pastel_comercio(labels, valores, nombre_archivo, colores):
            fig, ax = plt.subplots(figsize=SIZE_GRAFICO_PASTEL_PEQUENO)
            fig.subplots_adjust(left=0.08, right=0.92, top=0.92, bottom=0.08)
            
            wedges, texts, autotexts = ax.pie(
                valores, labels=None, colors=colores[:len(valores)],
                autopct=lambda p: f"{p:.2f}%", startangle=90, labeldistance=1.08, 
                pctdistance=0.72, textprops={'fontsize': FONT_SIZE_EJES}
            )
            
            for i, wedge in enumerate(wedges):
                angulo = (wedge.theta2 + wedge.theta1) / 2
                x = np.cos(np.deg2rad(angulo))
                y = np.sin(np.deg2rad(angulo))
                ax.text(x * 1.18, y * 1.18, labels[i], ha='center', va='center', fontsize=FONT_SIZE_EJES)
                
            ax.set_aspect('equal', adjustable='box')
            ruta = ASSETS_DIR / nombre_archivo
            plt.savefig(ruta, dpi=DPI_ESTANDAR, bbox_inches="tight", transparent=True)
            plt.close()
            return ruta

        # Datos Gráficos Página 3
        tabla_horarios_percepcion_df = df.iloc[335:344, [0,1,2]].copy().dropna(how="all")
        horarios_labels = tabla_horarios_percepcion_df.iloc[:,0].astype(str).tolist()
        horarios_porcentajes = (
            tabla_horarios_percepcion_df.iloc[:,1].astype(str).str.replace("%","").str.replace(",","")
        )
        horarios_porcentajes = pd.to_numeric(horarios_porcentajes, errors="coerce").fillna(0)
        
        tabla_horarios_percepcion = []
        for _, row in tabla_horarios_percepcion_df.iterrows():
            categoria = row.iloc[0]
            frecuencia = row.iloc[2]
            if pd.notna(categoria) and pd.notna(frecuencia):
                tabla_horarios_percepcion.append([str(categoria), int(frecuencia)])
                
        if not tabla_horarios_percepcion_df.empty:
            frecuencias = pd.to_numeric(tabla_horarios_percepcion_df.iloc[:,2], errors="coerce").fillna(0)
            idx_max = frecuencias.idxmax()
            horario_mayor = str(tabla_horarios_percepcion_df.loc[idx_max].iloc[0])
        else:
            horario_mayor = "No disponible"

        tabla_armas_df = df.iloc[349:357, [0,1,2]].copy().dropna(how="all")
        armas_labels = tabla_armas_df.iloc[:,0].astype(str).tolist()
        armas_porcentajes = tabla_armas_df.iloc[:,1].astype(str).str.replace("%","").str.replace(",","")
        armas_porcentajes = pd.to_numeric(armas_porcentajes, errors="coerce").fillna(0)
        
        tabla_armas = []
        for _, row in tabla_armas_df.iterrows():
            categoria = row.iloc[0]
            frecuencia = row.iloc[2]
            if pd.notna(categoria) and pd.notna(frecuencia):
                tabla_armas.append([str(categoria), int(frecuencia)])
                
        if not tabla_armas_df.empty:
            frecuencias_armas = pd.to_numeric(tabla_armas_df.iloc[:,2], errors="coerce").fillna(0)
            idx_max_armas = frecuencias_armas.idxmax()
            metodo_mas_usado = str(tabla_armas_df.loc[idx_max_armas].iloc[0])
        else:
            metodo_mas_usado = "No disponible"

        omitidas_aportes = ws["G322"].value
        grafico_horarios_percepcion = generar_grafico_horarios_percepcion(horarios_labels, horarios_porcentajes)
        grafico_armas_percepcion = generar_grafico_armas_percepcion(armas_labels, armas_porcentajes)

        # ================= SERVICIO POLICIAL Y COMERCIO =================
        labels_servicio = df.iloc[362:367,0].tolist()
        valores_servicio = (df.iloc[362:367,1].astype(float) * 100).tolist()
        tabla_servicio = list(zip(df.iloc[362:367,0], df.iloc[362:367,2]))

        labels_servicio_anual = df.iloc[372:375,0].tolist()
        valores_servicio_anual = df.iloc[372:375,1].astype(float).tolist()
        tabla_servicio_anual = list(zip(df.iloc[372:375,0], df.iloc[372:375,2]))

        labels_conoce = df.iloc[380:382,0].tolist()
        valores_conoce = df.iloc[380:382,1].astype(float).tolist()
        tabla_conoce = list(zip(df.iloc[380:382,0], df.iloc[380:382,2]))

        labels_conversado = df.iloc[387:389,0].tolist()
        valores_conversado = df.iloc[387:389,1].astype(float).tolist()
        tabla_conversado = list(zip(df.iloc[387:389,0], df.iloc[387:389,2]))

        omitidas_servicio = int(df.iloc[386,6])
        total_respuestas_servicio = int(df.iloc[382,2])

        # Sector Comercial
        labels_comercio_seguridad = [lbl.replace("Ni seguro ni inseguro", "Ni seguro\nni inseguro") for lbl in df.iloc[398:400,0].tolist()]
        valores_comercio_seguridad = df.iloc[398:400,1].astype(float).tolist()
        grafico_comercio_seguridad = generar_grafico_pastel_comercio(
            labels_comercio_seguridad, valores_comercio_seguridad, "grafico_comercio_seguridad.png", PALETA_COMERCIO_1
        )

        labels_comercio_programa = df.iloc[405:407,0].tolist()
        valores_comercio_programa = df.iloc[405:407,1].astype(float).tolist()
        grafico_comercio_programa = generar_grafico_pastel_comercio(
            labels_comercio_programa, valores_comercio_programa, "grafico_comercio_programa.png", PALETA_COMERCIO_2
        )

        labels_comercio_inscrito = df.iloc[412:414, 0].astype(str).str.strip().tolist()
        serie_comercio = df.iloc[412:414, 1]
        if serie_comercio.astype(str).str.contains("#DIV/0!").any():
            valores_comercio_inscrito = [100.0 if lbl.upper() == "NO" else 0.0 for lbl in labels_comercio_inscrito]
        else:
            valores_comercio_inscrito = pd.to_numeric(serie_comercio, errors="coerce").fillna(0).tolist()
            
        grafico_comercio_inscrito = generar_grafico_pastel_comercio(
            labels_comercio_inscrito, valores_comercio_inscrito, "grafico_comercio_inscrito.png", PALETA_COMERCIO_1
        )

        labels_comercio_contacto = df.iloc[419:421,0].tolist()
        valores_comercio_contacto = df.iloc[419:421,1].astype(float).tolist()
        grafico_comercio_contacto = generar_grafico_pastel_comercio(
            labels_comercio_contacto, valores_comercio_contacto, "grafico_comercio_contacto.png", PALETA_COMERCIO_2
        )

        def generar_grafico_servicio_policial(labels, valores):
            plt.figure(figsize=SIZE_GRAFICO_SERVICIO_POLICIAL)
            ax = plt.gca()
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_visible(False)
            ax.set_yticks([])
            ax.tick_params(left=False)
            
            plt.bar(labels, valores, color=PALETA_SERVICIO)
            for i, v in enumerate(valores):
                plt.text(i, v + 0.01, f"{v:.2f}%", ha="center", fontsize=FONT_SIZE_EJES)
                
            plt.ylim(0, max(valores) * 1.05)
            ruta = ASSETS_DIR / "grafico_servicio_policial.png"
            plt.savefig(ruta, dpi=DPI_ESTANDAR, bbox_inches="tight")
            plt.close()
            return ruta

        def generar_pie_servicio(labels, valores, nombre, texto_size=FONT_SIZE_EJES):
            plt.figure(figsize=SIZE_GRAFICO_PASTEL_PEQUENO)
            plt.pie(
                valores, labels=None, colors=PALETA_SERVICIO,
                autopct=lambda p: f"{p:.2f}%", pctdistance=0.7, textprops={'fontsize': texto_size}
            )
            plt.axis("equal")
            ruta = ASSETS_DIR / nombre
            plt.savefig(ruta, dpi=DPI_ESTANDAR, bbox_inches="tight")
            plt.close()
            return ruta

        grafico_servicio_policial = generar_grafico_servicio_policial(labels_servicio, valores_servicio)
        grafico_servicio_anual = generar_pie_servicio(labels_servicio_anual, valores_servicio_anual, "grafico_servicio_anual.png")
        grafico_conoce_policia = generar_pie_servicio(labels_conoce, valores_conoce, "grafico_conoce.png", texto_size=FONT_SIZE_TEXTO_DESTACADO)

        # ================= GRÁFICO ATENCIÓN (BARRAS HORIZONTALES) =================
        atencion_df = df.iloc[385:392, [0,1,2]].copy().dropna(how="all")
        atencion_df.columns = ["categoria", "porcentaje", "frecuencia"]
        atencion_df["porcentaje"] = atencion_df["porcentaje"].astype(str).str.replace("%","").str.replace(",",".")
        atencion_df["porcentaje"] = pd.to_numeric(atencion_df["porcentaje"], errors="coerce").fillna(0)
        
        if atencion_df["porcentaje"].max() <= 1:
            atencion_df["porcentaje"] = atencion_df["porcentaje"] * 100
            
        atencion_df["frecuencia"] = pd.to_numeric(atencion_df["frecuencia"], errors="coerce").fillna(0)
        
        labels_atencion = atencion_df["categoria"].astype(str).tolist()
        porcentajes_atencion = atencion_df["porcentaje"].tolist()
        frecuencias_atencion = atencion_df["frecuencia"].astype(int).tolist()

        tabla_atencion = []
        for l, p, f in zip(labels_atencion, porcentajes_atencion, frecuencias_atencion):
            tabla_atencion.append([l, str(f)])

        def generar_grafico_atencion():
            fig, ax = plt.subplots(figsize=SIZE_GRAFICO_ATENCION)
            # Usando un azul de la paleta principal
            barras = ax.barh(labels_atencion, porcentajes_atencion, color=PALETA_COMPLETA[2])
            ax.set_xlim(0, 100)
            
            for bar, valor in zip(barras, porcentajes_atencion):
                ax.text(
                    valor + 1, bar.get_y() + bar.get_height()/2,
                    f"{valor:.2f}%", va="center", fontsize=FONT_SIZE_BARRAS_VALORES, color=COLOR_TEXTO
                )
                
            for spine in ax.spines.values():
                spine.set_visible(False)
            ax.tick_params(left=False, bottom=False)
            plt.tight_layout()
            ruta = ASSETS_DIR / "grafico_atencion.png"
            plt.savefig(ruta, dpi=DPI_ESTANDAR, bbox_inches="tight", transparent=True)
            plt.close()
            return ruta

        grafico_atencion = generar_grafico_atencion()

        # ================= GENERAR PDF =================
        if st.button("HACER INFORME TERRITORIAL"):
            pdf_buffer = generar_pdf(
                portada_path=str(ASSETS_DIR / "portada.png"),
                grafico_relacion_path=str(grafico_rel_path),
                grafico_edad_path=str(grafico_edad_path),
                grafico_escolaridad_path=str(grafico_escolaridad_path),
                grafico_genero_path=str(grafico_genero_path),
                delegacion=delegacion,
                codigo=codigo,
                tabla_participacion=tabla_participacion,
                tabla_edad=tabla_edad,
                tabla_escolaridad=tabla_escolaridad,
                tabla_genero=tabla_genero,
                tabla_encuesta_comunidad=tabla_encuesta_comunidad,
                tabla_otras_encuestas=tabla_otras_encuestas,
                datos_pagina_8=datos_pagina_8,
                datos_pagina_9=datos_pagina_9,
                tabla_delitos=tabla_delitos,
                tabla_riesgos=tabla_riesgos,
                porcentaje_delitos=porcentaje_delitos,
                porcentaje_riesgos=porcentaje_riesgos,
                cantidad_delitos=cantidad_delitos,
                cantidad_riesgos=cantidad_riesgos,
                micmac_poder=micmac_poder,
                micmac_conflicto=micmac_conflicto,
                micmac_autonomas=micmac_autonomas,
                micmac_resultados=micmac_resultados,
                tabla_riesgos_micmac2=tabla_riesgos_micmac2,
                tabla_delitos_micmac2=tabla_delitos_micmac2,
                cantidad_problematicas=cantidad_problematicas,
                riesgos_total=riesgos_total,
                delitos_total=delitos_total,
                causas_identificadas=causas_identificadas,
                factores_micmac=factores_micmac,
                triangulo_directa=triangulo_directa,
                triangulo_sociocultural=triangulo_sociocultural,
                triangulo_estructural=triangulo_estructural,
                tabla_instituciones=tabla_instituciones,
                grafico_denuncias_path="assets/grafico_denuncias.png",
                tabla_denuncias=tabla_denuncias,
                total_denuncias=total_denuncias,
                grafico_horario_path=str(ASSETS_DIR / "grafico_horario.png"),
                tabla_horario=tabla_horario,
                total_am=total_am,
                total_pm=total_pm,
                tabla_horario_distrito=tabla_horario_distrito,
                grafico_p14_path=str(ASSETS_DIR / "grafico_p14.png"),
                tabla_p14=tabla_p14,
                grafico_p15_path=str(ASSETS_DIR / "grafico_p15.png"),
                tabla_p15=tabla_p15,
                total_lineas=total_lineas,
                lineas_municipalidad=lineas_municipalidad,
                lineas_fp=lineas_fp,
                lineas_mixtas=lineas_mixtas,
                logo_muni_path=str(logo_muni_path),
                lineas_accion_data=lineas_accion_data,
                grafico_percepcion_actual_path=str(ASSETS_DIR / "grafico_percepcion_actual.png"),
                grafico_percepcion_comparacion_path=str(ASSETS_DIR / "grafico_percepcion_comparacion.png"),
                tabla_percepcion=tabla_percepcion,
                grafico_victimizacion_path=str(ASSETS_DIR / "grafico_victimizacion.png"),
                grafico_no_denuncia_path=str(ASSETS_DIR / "grafico_no_denuncia.png"),
                tabla_no_denuncia=tabla_no_denuncia,
                motivo_principal=motivo_principal,
                total_omitidas=total_omitidas,
                grafico_horarios_percepcion=grafico_horarios_percepcion,
                grafico_armas_percepcion=grafico_armas_percepcion,
                tabla_horarios_percepcion=tabla_horarios_percepcion,
                tabla_armas=tabla_armas,
                horario_mayor=horario_mayor,
                metodo_mas_usado=metodo_mas_usado,
                omitidas_aportes=omitidas_aportes,
                grafico_servicio_policial=grafico_servicio_policial,
                grafico_servicio_anual=grafico_servicio_anual,
                grafico_conoce_policia=grafico_conoce_policia,
                grafico_atencion=grafico_atencion,
                tabla_servicio=tabla_servicio,
                tabla_servicio_anual=tabla_servicio_anual,
                tabla_conoce=tabla_conoce,
                tabla_atencion=tabla_atencion,
                omitidas_servicio=omitidas_servicio,
                total_respuestas_servicio=total_respuestas_servicio,
                grafico_comercio_seguridad=grafico_comercio_seguridad,
                grafico_comercio_programa=grafico_comercio_programa,
                grafico_comercio_inscrito=grafico_comercio_inscrito,
                grafico_comercio_contacto=grafico_comercio_contacto,
            )
            
            pdf_bytes = pdf_buffer.getvalue()
            st.subheader("Informe generado correctamente")
            st.download_button(
                label="Descargar PDF",
                data=pdf_bytes,
                file_name="informe.pdf",
                mime="application/pdf"
            )

    except Exception as e:
        import traceback
        st.error("ERROR DETALLADO:")
        st.code(traceback.format_exc())
